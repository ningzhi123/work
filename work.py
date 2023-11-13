import openpyxl
import sys

# 打开Excel文件
wb = openpyxl.load_workbook(sys.argv[1])

# 获取某一个sheet
# sheet = wb.get_sheet_by_name(sys.argv[2])
sheet = wb[sys.argv[2]]

# 获取sheet中的某一个单元格的值
for i in range(1, sheet.max_row):
    cell_value = sheet.cell(row=i, column=4).value

    if cell_value == sys.argv[3]:
        # 输出单元格的值
        print(cell_value)
        print(sheet.cell(row=i, column=3).value)
        # 新建excel文件
        wb1 = openpyxl.Workbook()
        wb1_sheet = wb1.active
        wb1_sheet.column_dimensions['A'].width = 14
        wb1_sheet.column_dimensions['B'].width = 10
        wb1_sheet.column_dimensions['C'].width = 14
        wb1_sheet.column_dimensions['D'].width = 6
        wb1_name = cell_value+"_" + \
            str(sheet.cell(row=3, column=6).value)[5:10]+".xlsx"
        wb1_row = 0
        # 新建txt文档
        txt_name = cell_value+"_" + \
            str(sheet.cell(row=3, column=6).value)[5:10]+".txt"
        f = open(txt_name, "w")

        for j in range(5, sheet.max_column + 1):
            wb1_row += 1
            if '星期' in sheet.cell(row=1, column=j).value:
                txt = ""
                if "-" in str(sheet.cell(row=3, column=j).value):
                    txt = str(sheet.cell(row=3, column=j).value)[5:10].center(6) + str(sheet.cell(row=2, column=j).value).center(
                        3) + str(sheet.cell(row=i, column=j).value).center(7) + str(sheet.cell(row=i+1, column=j).value).center(5)
                    print(txt)
                    f.write(txt)
                    f.write("\r\n")
                else:
                    txt = str(sheet.cell(row=3, column=j).value).center(6) + str(sheet.cell(row=2, column=j).value).center(
                        3) + str(sheet.cell(row=i, column=j).value).center(7) + str(sheet.cell(row=i+1, column=j).value).center(5)
                    print(txt)
                    f.write(txt)
                    f.write("\r\n")
                wb1_sheet.cell(row=wb1_row, column=1).value = sheet.cell(
                    row=3, column=j).value
                wb1_sheet.cell(row=wb1_row, column=1).alignment = openpyxl.styles.Alignment(
                    horizontal='center')
                wb1_sheet.cell(row=wb1_row, column=1).number_format = 'yyyy-mm-dd;@'

                wb1_sheet.cell(row=wb1_row, column=2).value = sheet.cell(
                    row=2, column=j).value
                wb1_sheet.cell(row=wb1_row, column=2).alignment = openpyxl.styles.Alignment(
                    horizontal='center')

                wb1_sheet.cell(row=wb1_row, column=3).value = sheet.cell(
                    row=i, column=j).value
                wb1_sheet.cell(row=wb1_row, column=3).alignment = openpyxl.styles.Alignment(
                    horizontal='center')

                wb1_sheet.cell(row=wb1_row, column=4).value = sheet.cell(
                    row=i+1, column=j).value
                wb1_sheet.cell(row=wb1_row, column=4).alignment = openpyxl.styles.Alignment(
                    horizontal='center')
                wb1.save(wb1_name)
            else:
                txt = str(sheet.cell(row=1, column=j).value).ljust(
                    8) + ":" + str(sheet.cell(row=i, column=j).value).ljust(8)
                print(txt)
                f.write(txt)
                f.write("\r\n")
                wb1_sheet.cell(row=wb1_row, column=1).value = sheet.cell(
                    row=2, column=j).value
                wb1_sheet.cell(row=wb1_row, column=1).alignment = openpyxl.styles.Alignment(
                    horizontal='center')
                wb1_sheet.cell(row=wb1_row, column=2).value = sheet.cell(
                    row=i, column=j).value
                wb1_sheet.cell(row=wb1_row, column=2).alignment = openpyxl.styles.Alignment(
                    horizontal='center')
                wb1.save(wb1_name)

        wb1.save(wb1_name)
        f.close()
